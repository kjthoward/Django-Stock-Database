from dateutil.relativedelta import relativedelta
from django.http import HttpResponse, HttpResponseRedirect
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.models import User, Group
from django.db.models import F, Q
from django.shortcuts import render
from django.urls import reverse
from django.db import transaction
from operator import attrgetter
import openpyxl
import pdb
import datetime
import math
import random
import string
from .prime import PRIME
from .email import send, EMAIL
from .pdf_report import report_gen
from .models import ForceReset, Suppliers, Teams, Reagents, Internal, Validation, Recipe, Inventory, Solutions, VolUsage
from .forms import LoginForm, NewInvForm1, NewInvForm, NewProbeForm, UseItemForm, OpenItemForm, ValItemForm, FinishItemForm,\
                   NewSupForm, NewTeamForm, NewReagentForm, NewRecipeForm, SearchForm, ChangeDefSupForm1, ChangeDefSupForm, ChangeDefTeamForm1, \
                   ChangeDefTeamForm, RemoveSupForm, EditSupForm, EditTeamForm, EditReagForm, EditInvForm, DeleteForm, UnValForm, ChangeUseForm, \
                   ChangeMinForm1, ChangeMinForm, InvReportForm,StockReportForm, PWResetForm, PasswordChangeForm, WitnessForm, TeamOnlyForm, ValeDatesForm

LOGINURL = settings.LOGIN_URL
RESETURL = "/stock/forcereset/"
UNAUTHURL = "/stock/unauth/"

CONDITIONS={"GD":"Good",
            "DU":"Damaged - Usable",
            "UN":"Damaged - Not Usable",
            "NA":"N/A, Made in House"}

#used in user_passes_test decorator to check if someone if logged in
def is_logged_in(user):
    return user.is_active
#used in user_passes_test decorator to check if the account logged in is admin
def is_admin(user):
    return (user.is_staff or user.groups.filter(name="Admin").exists())

def is_super_admin(user):
    return user.is_staff
#used in user_passes_test decorator to check if the account has a forced password reset active (decorate used as
#even though after logging in with a reset password it prompts you to change, could go to any link manually to skip
#decorator means you will always be brought back to change password
def no_reset(user):
    if ForceReset.objects.get(user=user.pk).force_password_change==True:
        return False
    else:
        return True

def view_404(httprequest, exception):
    messages.success(httprequest, f"Webpage with the path {exception.args[-1]['path']} doesn't exist.\n If you think you are getting this message in error please contact a member of the Bioinformatics Team")
    return HttpResponseRedirect(reverse("stock_web:listinv"))

#sets up database for first use (sets up groups and adds Admin as superuser/staff)
def prime(httprequest):
    if len(User.objects.all())!=0:
        messages.success(httprequest, "Database has already been primed!")
        return HttpResponseRedirect(reverse("stock_web:listinv"))
    else:
        PRIME()
        messages.success(httprequest, "Database Primed")
        return HttpResponseRedirect(reverse("stock_web:listinv"))

def _toolbar(httprequest, active=""):
    inventory_dropdown = [{"name":"All", "url":reverse("stock_web:inventory", args=["_", "all","_",1])},
                          {"name":"In Stock", "url":reverse("stock_web:inventory", args=["_", "instock","_", 1])},
                          {"name":"Expiring Soon", "url":reverse("stock_web:inventory", args=["_", "expsoon","_",1])},
                          {"name":"Solutions", "url":reverse("stock_web:inventory", args=["_", "solutions","_",1])},
                          {"name":"Validated", "url":reverse("stock_web:inventory", args=["_", "validated","_",1])},
                          {"name":"Not validated", "url":reverse("stock_web:inventory", args=["_", "notvalidated","_",1])},
                          {"name":"List View","url":reverse("stock_web:listinv")},
                          ]

    toolbar = [([{"name":"Inventory", "dropdown":inventory_dropdown},
                 {"name":"Recipes", "url":reverse("stock_web:recipes"), "glyphicon":"folder-open"},
                 {"name":"Stock Reports", "url":reverse("stock_web:stockreport", args=["_","_","_","_"]),"glyphicon":"download"},
                 ], "left")]



    undo_dropdown = [{"name": "Change Default Supplier", "url":reverse("stock_web:changedefsup", args=["_"])},
                     {"name": "Change Default Team", "url":reverse("stock_web:changedefteam", args=["_"])},
                     {"name": "Edit Minimum Stock Levels", "url":reverse("stock_web:changemin",args=["_"])},
                     {"name": "(De)Activate Reagents/Recipies", "url":reverse("stock_web:activreag")},
                     {"name": "(De)Activate Suppliers", "url":reverse("stock_web:activsup")},
                     {"name": "(De)Activate Team", "url":reverse("stock_web:activteam")},
                     {"name": "Remove Suppliers", "url":reverse("stock_web:removesup")},
                     {"name": "Edit Inventory Item", "url":reverse("stock_web:editinv", args=["_"])}]



    if is_admin(httprequest.user):
        toolbar[0][0].append({"name":"Inventory Reports", "url":reverse("stock_web:invreport", args=["_","_","_","_"]), "glyphicon":"list"})
        toolbar[0][0].append({"name":"Edit Data", "dropdown":undo_dropdown, "glyphicon":"wrench"})
        if httprequest.user.is_staff:
            toolbar[0][0].append({"name":"Update Users", "url":"/stock/admin/auth/user/","glyphicon":"user"})
        new_dropdown = [{"name": "Inventory Item", "url":reverse("stock_web:newinv", args=["_"])},
                        {"name":"Supplier", "url":reverse("stock_web:newsup")},
                        {"name":"Team", "url":reverse("stock_web:newteam")},
                        {"name":"Reagent", "url":reverse("stock_web:newreagent")},
                        {"name":"Recipe", "url":reverse("stock_web:newrecipe")}]
        toolbar.append(([{"name": "new", "glyphicon": "plus", "dropdown": new_dropdown}],"right"))
        search_dropdown = [{"name": "Search", "url": reverse("stock_web:search")},
                           {"name": "Validation Dates", "url": reverse("stock_web:valdates")}]
        toolbar[1][0].append({"name": "Search", "glyphicon":"search", "dropdown": search_dropdown})

    else:
        toolbar.append(([{"name": "New Inventory Item", "glyphicon": "plus", "url":reverse("stock_web:newinv", args=["_"])}],"right"))
        toolbar[1][0].append({"name": "Search", "glyphicon": "search", "url": reverse("stock_web:search")})

    toolbar[1][0].append({"name":"Account Settings", "glyphicon":"cog", "dropdown":[
                         {"name": "Logout "+str(httprequest.user), "url": reverse("stock_web:loginview")},
                         {"name":"Change Password", "url":reverse("stock_web:change_password")}]})

    for entry in toolbar[0][0]:
        if entry["name"] == active:
            entry["active"] = True
    for entry in toolbar[1][0]:
        if entry["name"] == active:
            entry["active"] = True
    return toolbar

@user_passes_test(is_logged_in, login_url=LOGINURL)
def change_password(httprequest):
    if httprequest.method == 'POST':
        form = PasswordChangeForm(httprequest.user, httprequest.POST)
        if form.is_valid():
            if form.cleaned_data["old_password"]==form.cleaned_data["new_password1"]:
                messages.success(httprequest, 'Your new password cannot be your old password')
                return HttpResponseRedirect(reverse("stock_web:change_password"))
            user = form.save()
            update_session_auth_hash(httprequest, user)  # Important!
            messages.success(httprequest, 'Your password was successfully updated!')
            try:
                if ForceReset.objects.get(user=httprequest.user.pk).force_password_change==True:
                    reset=ForceReset.objects.get(user_id=user.id)
                    reset.force_password_change=False
                    reset.save()
            except:
                pass
            return HttpResponseRedirect(reverse("stock_web:listinv"))
        else:
            errors=[]
            for v in form.errors.values():
                for error in v:
                    errors+=[error.replace("didn't", "do not")]
            messages.success(httprequest, (" \n".join(errors)))
            form.errors.clear()
            messages.info(httprequest, "Your password can't be too similar to your other personal information.")
            messages.info(httprequest, "Your password must contain at least 8 characters.")
            messages.info(httprequest, "Your password can't be a commonly used password.")
            messages.info(httprequest, "Your password can't be entirely numeric.")
    else:
        form = PasswordChangeForm(httprequest.user)
        messages.info(httprequest, "Your password can't be too similar to your other personal information.")
        messages.info(httprequest, "Your password must contain at least 8 characters.")
        messages.info(httprequest, "Your password can't be a commonly used password.")
        messages.info(httprequest, "Your password can't be entirely numeric.")
    submiturl = reverse("stock_web:change_password")
    cancelurl = reverse("stock_web:listinv")
    context={"form": form, "heading":"Change Password for {}".format(httprequest.user),
             "submiturl": submiturl, "cancelurl": cancelurl, "toolbar":_toolbar(httprequest, active="Account Settings")}
##    if ForceReset.objects.get(user=httprequest.user.pk).force_password_change==False:
##        context.update({"toolbar":_toolbar(httprequest)})
    return render(httprequest, "stock_web/pwform.html", context)

def resetpw(httprequest):
    if EMAIL==False:
        messages.success(httprequest, "Email resets are not currently enabled. Contact an Admin to reset your password")
        return HttpResponseRedirect(reverse("stock_web:loginview"))
    if httprequest.method == 'POST':
        form=PWResetForm(httprequest.POST)

        if form.is_valid():
            new_pw=''.join(random.choice(string.ascii_letters + string.digits) for _ in range(15))
            USER = User.objects.get(username=form.data["user"])
            USER.set_password(new_pw)
            USER.save()
            try:
                reset=ForceReset.objects.get(user_id=USER.id)
                reset.force_password_change=True
            #except just incase user somehow doesn't exist in force reset table (but they should be synced)
            except:
                reset=ForceReset.objects.create(user=USER, force_password_change=True)
            reset.save()
            subject="Password for stock database account '{}' has been reset.".format(USER.username)
            text="<p>The Password for your stock database account '{}' has been reset.<br><br>".format(USER.username)
            text+="Your new password is: {}<br><br>".format(new_pw)
            text+="You will be required to change this when you login.<br><br>"
            try:
                send(subject,text,USER.email)
                messages.success(httprequest, "Password has been reset and emailed to {}".format(USER.email))
            except:
                messages.success(httprequest, "Email did not send. Please try again. If error persists contact an Admin")
            return HttpResponseRedirect(reverse("stock_web:loginview"))
    else:
        form=PWResetForm()
    submiturl = reverse("stock_web:resetpw")
    cancelurl = reverse("stock_web:listinv")
    return render(httprequest, "stock_web/pwform.html", {"form": form, "heading":"Enter Username to Reset Password for", "submiturl": submiturl, "cancelurl": cancelurl })

def forcereset(httprequest):
    messages.success(httprequest, "You are required to change your password after resetting it")
    return HttpResponseRedirect(reverse("stock_web:change_password"))

def unauth(httprequest):
    messages.success(httprequest, "The page you are trying to access requires Admin rights.\
                     Please login using an Admin account.")
    return HttpResponseRedirect(reverse("stock_web:listinv"))

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def search(httprequest):
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "search":
            return HttpResponseRedirect(reverse("stock_web:listinv"))
        else:
            form = SearchForm(httprequest.POST)
            if form.is_valid():
                queries = []
                for key, query in [("reagent", "reagent__name__icontains"), ("supplier", "supplier__name__icontains"),
                                   ("lot_no", "lot_no__icontains"), ("int_id","internal__batch_number__exact"),
                                   ("val_status", "val_id__isnull"),("in_stock","finished__lte"),( "rec_range","date_rec__range"), 
                                   ("open_range","date_op__range"), ("val_range","val_id__val_date__range"), 
                                   ("fin_range","date_fin__range"), ("team", "team__exact"), ("inc_open","is_op__lte")
                                  ]:
                    val = form.cleaned_data[key]
                    if key=="team" and val is not None:
                        val=str(val.id)
                    if val:
                        if val[0]!=None:
                            if "range" in query:
                                val=(val[0].strftime("%Y-%m-%d"), val[1].strftime("%Y-%m-%d"))
                            if key=="in_stock" and val=="2":
                                query="finished__exact"
                                val="1"
                            queries += ["{}={}".format(query, val)]
                return HttpResponseRedirect(reverse("stock_web:inventory", args=["search", ";".join(queries),"_","1"]))
    else:
        form = SearchForm(initial = {"in_stock":1, "inc_open":1})
    submiturl = reverse("stock_web:search")
    cancelurl = reverse("stock_web:listinv")
    return render(httprequest, "stock_web/searchform.html", {"form": form, "heading":"Enter Search Query", "toolbar": _toolbar(httprequest, active="Search"), "submiturl": submiturl, "cancelurl": cancelurl })


def loginview(httprequest):
    if httprequest.user.is_authenticated:
        logout(httprequest)
        messages.success(httprequest,"You are now logged out")
        return HttpResponseRedirect(reverse("stock_web:listinv"))
    else:

        if httprequest.method == "POST" and "login" in httprequest.POST:
            form = LoginForm(httprequest.POST)

            if form.is_valid():
                user = authenticate(username=form.cleaned_data["username"], password=form.cleaned_data["password"])
                if user is not None and user.is_active:
                    login(httprequest, user)
                    #Autosets the staff status based on group
                    if user.groups.filter(name="Senior_Admin").exists() or user.is_superuser:
                        user.is_staff=True
                    else:
                        user.is_staff=False
                    user.save()
                    if ForceReset.objects.get(user=httprequest.user.pk).force_password_change==True:
                        messages.success(httprequest,"You are required to change your password after resetting it")
                        return HttpResponseRedirect(reverse("stock_web:change_password"))
                    else:
                        return HttpResponseRedirect(httprequest.GET["next"] if "next" in httprequest.GET.keys() else reverse("stock_web:listinv"))
                else:
                    try:
                        User.objects.get(username=form.cleaned_data["username"])
                        messages.success(httprequest, "Incorrect password entered")
                    except:
                        messages.success(httprequest, f'User {form.cleaned_data["username"]} does not exist')
            else:
                pdb.set_trace()

        else:
            form = LoginForm()
            if "rv:11.0" in httprequest.META['HTTP_USER_AGENT']:
                messages.success(httprequest, "WARNING - YOUR BROWSER IS NOT SUPPORTED ON THIS SITE. PLEASE SWITCH TO FIREFOX OR GOOGLE CHROME")
    return render(httprequest, "stock_web/login.html", {"form": form})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def valdates(httprequest):
    submiturl = reverse("stock_web:valdates")
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Search")
    header = "Select Date Range to Search for"
    form=ValeDatesForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or "Download" not in httprequest.POST["submit"]:
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():
                start, end = form.cleaned_data["val_range"]
                title="Items validated between {} and {} - Downloaded {}".format(start.strftime("%d-%m-%Y"), end.strftime("%d-%m-%Y"), datetime.datetime.today().date().strftime("%d-%m-%Y"))
                items=Inventory.objects.select_related("supplier","reagent","internal","val", "op_user").filter(val_id__val_date__range=[start, end])
                body=[["Reagent", "Catalogue Number", "Supplier", "Lot Number", "Stock Number", "Received",
                       "Expiry", "Opened", "Opened By", "Date Validated", "Validation Run"]]
                for item in items:
                    body+= [[item.reagent.name,
                              item.reagent.cat_no,
                              item.supplier.name,
                              item.lot_no,
                              item.internal.batch_number,
                              item.date_rec.strftime("%d/%m/%Y"),
                              item.date_exp.strftime("%d/%m/%Y"),
                              item.date_op.strftime("%d/%m/%Y") if item.date_op is not None else "",
                              item.op_user.username if item.op_user is not None else "",
                              item.val.val_date.strftime("%d/%m/%Y") if item.val is not None else "",
                              item.val.val_run if item.val is not None else "",
                              ]]
                if "pdf" in httprequest.POST["submit"]:
                    httpresponse = HttpResponse(content_type='application/pdf')
                    httpresponse['Content-Disposition'] = 'attachment; filename="{}.pdf"'.format(title)
                    table=report_gen(body,title,httpresponse,httprequest.user.username)
                elif "xlsx" in httprequest.POST["submit"]:
                    workbook = openpyxl.Workbook()
                    worksheet = workbook.active
                    for row in body:
                        worksheet.append(row)
                    httpresponse = HttpResponse(content=openpyxl.writer.excel.save_virtual_workbook(workbook), content_type='application/ms-excel')
                    httpresponse['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(title)
                return httpresponse
    else:
        form = form()
        return render(httprequest, "stock_web/reportform.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def listinv(httprequest):
    title = "List of Reagents"
    headings = ["Reagent Name", "Number In Stock", "Minimum Stock Level"]
    items=Reagents.objects.all().exclude(is_active=False, count_no=0).order_by("name")
    body=[]
    for item in items:
        values = [item.name,
                  "{}µl".format(item.count_no) if item.track_vol==True else item.count_no,
                  "{}µl".format(item.min_count) if item.track_vol==True else item.min_count]
        urls=[reverse("stock_web:inventory",args=["filter","reagent__name__iexact={};finished__lte=0".format(item.name),"_",1]),
              "",
              "",
              ]
        body.append((zip(values,urls), True if item.count_no<item.min_count else False))

    context = {"header":title,"headings":headings, "body":body, "toolbar":_toolbar(httprequest, active="Inventory")}
    return render(httprequest, "stock_web/list.html", context)

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def inventory(httprequest, search, what, sortby, page):
    #forces page 1 if non numberical value entered or <1
    try:
        page=int(page)
    except:
        page=1
    if page<1:
        page=1
    ##HACK USED IF / IS IN REAGENT SEARCH TERM (such as "item1/item2 mix")
    if "/" in search and (("__icontains" in what) or ("__lte" in what) or ("__iexact" in what)):
        what="".join(search.split("/")[1:])+"/"+what
        search="search"
    if sortby!="_":
        sortquery = sortby.split("=")[1]
        if sortquery=="days_rem":
            pos=True
            sortby="_"
        elif sortquery=="-days_rem":
            pos=False
            sortby="_"
    else:
        sortquery="_"
    if what=="all":
        title = "Inventory - All Items"
        if sortby!="_":
            items=Inventory.objects.select_related("supplier","reagent","internal","val", "team").all().order_by(sortquery)
        else:
            items=Inventory.objects.select_related("supplier","reagent","internal","val", "team").all()
    if what=="instock":
        title = "Inventory - Items In Stock"
        if sortby!="_":
            items=Inventory.objects.select_related("supplier","reagent","internal","val", "team").filter(finished=False).order_by(sortquery)
        else:
            items=Inventory.objects.select_related("supplier","reagent","internal","val", "team").filter(finished=False)
    elif what=="solutions":
        title = "Inventory - Solutions"
        if sortby!="_":
            items=Inventory.objects.select_related("supplier","reagent","internal","val", "team").filter(sol_id__isnull=False, finished=False).order_by(sortquery)
        else:
            items=Inventory.objects.select_related("supplier","reagent","internal","val", "team").filter(sol_id__isnull=False, finished=False)
    elif what=="validated":
        title = "Inventory - Validated Items"
        if sortby!="_":
            items=Inventory.objects.select_related("supplier","reagent","internal","val", "team").filter(val_id__isnull=False, finished=False).order_by(sortquery)
        else:
            items=Inventory.objects.select_related("supplier","reagent","internal","val", "team").filter(val_id__isnull=False, finished=False)
    elif what=="notvalidated":
        title = "Inventory - Items Not Validated"
        if sortby!="_":
            items=Inventory.objects.select_related("supplier","reagent","internal","val", "team").filter(val_id__isnull=True, finished=False).order_by(sortquery)
        else:
            items=Inventory.objects.select_related("supplier","reagent","internal","val", "team").filter(val_id__isnull=True, finished=False)
    elif what=="expsoon":
        title = "Inventory - Items Expiring Within 6 Weeks"
        if sortby!="_":
            items=Inventory.objects.select_related("supplier","reagent","internal","val", "team").filter(date_exp__lte=datetime.datetime.now()+datetime.timedelta(days=42), finished=False).order_by(sortquery)
        else:
            items=Inventory.objects.select_related("supplier","reagent","internal","val", "team").filter(date_exp__lte=datetime.datetime.now()+datetime.timedelta(days=42), finished=False)

    elif search=="search" or search=="filter":
        query = dict([q.split("=") for q in what.split(";")])
        if search=="search":
            for key, value in query.items():
                if "range" in key:
                    query[key]=value.strip("()").replace("'","").replace(" ","").split(",")
                if key=="val_id__isnull":
                    if int(value)==0:
                        value=False
                    else:
                        value=True
                    
                    query[key]=value

            title="Search Results"
            
        elif search=="filter" and "reagent__name__iexact" in query.keys():
            title=query['reagent__name__iexact']
        if sortby!="_":
            items=Inventory.objects.select_related("supplier","reagent","internal","val", "team").filter(**query).order_by(sortquery)
        else:
            items=Inventory.objects.select_related("supplier","reagent","internal","val", "team").filter(**query)
        if len(items)==1:
            return HttpResponseRedirect(reverse("stock_web:item",args=[items[0].id]))
    pages=[]
    if len(items)>200:
        for i in range(1, math.ceil(len(items)/200)+1):
            pages+=[[i,reverse("stock_web:inventory", args=[search, what, sortby, i]) if page!=i else ""]]
        #forces go to page 1 if number>last page manually entered
        if page>pages[-1][0]:
             return HttpResponseRedirect(reverse("stock_web:inventory", args=[search, what, sortby, 1]))
    headings = ["Reagent Name", "Catalogue Number", "Supplier", "Batch ID", "Date Received", "Expiry Date", "Date Opened", "Date Validated", "Date Finished","Days till expired", "Team"]
    headurls = [reverse("stock_web:inventory", args=[search, what,"order=-reagent_id__name"
                                                     if sortby=="order=reagent_id__name" else "order=reagent_id__name", 1]),
                reverse("stock_web:inventory", args=[search, what,"order=-reagent_id__cat_no"
                                                     if sortby=="order=reagent_id__cat_no" else "order=reagent_id__cat_no", 1]),
                reverse("stock_web:inventory", args=[search, what,"order=-supplier_id__name"
                                                     if sortby=="order=supplier_id__name" else "order=supplier_id__name", 1]),                                                     
                reverse("stock_web:inventory", args=[search, what,"order=-internal_id__batch_number"
                                                     if sortby=="order=internal_id__batch_number" else "order=internal_id__batch_number", 1]),
                reverse("stock_web:inventory", args=[search, what,"order=-date_rec"
                                                     if sortby=="order=date_rec" else "order=date_rec", 1]),
                reverse("stock_web:inventory", args=[search, what,"order=-date_exp"
                                                     if sortby=="order=date_exp" else "order=date_exp", 1]),
                reverse("stock_web:inventory", args=[search, what,"order=-date_op"
                                                     if sortby=="order=date_op" else "order=date_op",1]),
                reverse("stock_web:inventory", args=[search, what,"order=-val_id__val_date"
                                                     if sortby=="order=val_id__val_date" else "order=val_id__val_date",1]),
                reverse("stock_web:inventory", args=[search, what,"order=-date_fin"
                                                     if sortby=="order=date_fin" else "order=date_fin",1]),
                reverse("stock_web:inventory", args=[search, what,"order=-days_rem"
                                                     if sortquery=="days_rem" else "order=days_rem",1]),
                reverse("stock_web:inventory", args=[search, what,"order=-team"
                                                     if sortquery=="team" else "order=team",1])]
    headings=zip(headings,headurls)
    body=[]

    if "date_op" in sortby:
        q=items.extra(select={'date_op_null': 'date_op is null'})
        items = q.extra(order_by=['date_op_null',sortquery])
        ####################################################################
    if "days_rem" in sortquery:
        if pos==True:
            items = sorted(items, key = lambda item:int(item.days_remaining()), reverse=False)
        elif pos==False:
            items = sorted(items, key = lambda item:int(item.days_remaining()), reverse=True)



    items_trunc=items[(page-1)*200:page*200]
    for item in items_trunc:
        values = [item.reagent.name,
                  item.reagent.cat_no if item.reagent.cat_no is not None else "",
                  item.supplier.name,
                  item.internal.batch_number,
                  item.date_rec.strftime("%d/%m/%Y"),
                  item.date_exp.strftime("%d/%m/%Y"),
                  item.date_op.strftime("%d/%m/%Y") if item.date_op is not None else "",
                  item.val.val_date.strftime("%d/%m/%Y") if item.val_id is not None else "",
                  item.date_fin.strftime("%d/%m/%Y") if item.date_fin is not None else "",
                  item.days_remaining(),
                  item.team.name if item.team is not None else "",
                  ]
        urls=[reverse("stock_web:item",args=[item.id]),
              "",
              "",
              "",
              "",
              "",
              "",
              "",
              "",
              "",
              "",
              ]
        body.append((zip(values,urls),item.finished, True if item.days_remaining()<0 else False))
    context = {"header":title,"headings":headings, "body":body,
               "toolbar":_toolbar(httprequest, active="Inventory")}
    if pages:
        context.update({"pages":pages, "text1": "Click to change page",
                        "text2":"Current page is {} showing items {}-{}".format(page,(page-1)*200,
                                                                         page*200 if (page*200<len(items)) else (len(items)))})
    if httprequest.method=="POST" and search=="search":
        if ".xlsx" in httprequest.POST["submit"]:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            # import pdb; pdb.set_trace()
            worksheet.append([heading[0] for heading in headings])
            for item in items:
                worksheet.append([item.reagent.name,
                          item.reagent.cat_no,
                          item.supplier.name,
                          item.internal.batch_number,
                          item.date_rec.strftime("%d/%m/%Y"),
                          item.date_exp.strftime("%d/%m/%Y"),
                          item.date_op.strftime("%d/%m/%Y") if item.date_op is not None else "",
                          item.val.val_date.strftime("%d/%m/%Y") if item.val_id is not None else "",
                          item.date_fin.strftime("%d/%m/%Y") if item.date_fin is not None else "",
                          item.days_remaining(),
                          item.team.name if item.team is not None else "",
                          ])

            httpresponse = HttpResponse(content=openpyxl.writer.excel.save_virtual_workbook(workbook), content_type='application/ms-excel')
            httpresponse['Content-Disposition'] = 'attachment; filename="Search Results - {}.xlsx"'.format(str(datetime.datetime.today().strftime("%d/%m/%Y")))

        elif "pdf" in httprequest.POST["submit"]:
            doc_title="Search Results - {}".format(str(datetime.datetime.today().strftime("%d/%m/%Y")))
            contents=[[heading[0] for heading in headings]]
            for item in items:
                contents.append([item.reagent.name,
                          item.reagent.cat_no,
                          item.supplier.name,
                          item.internal.batch_number,
                          item.date_rec.strftime("%d/%m/%Y"),
                          item.date_exp.strftime("%d/%m/%Y"),
                          item.date_op.strftime("%d/%m/%Y") if item.date_op is not None else "",
                          item.val.val_date.strftime("%d/%m/%Y") if item.val_id is not None else "",
                          item.date_fin.strftime("%d/%m/%Y") if item.date_fin is not None else "",
                          item.days_remaining(),
                          item.team.name if item.team is not None else "",
                          ])
            httpresponse = HttpResponse(content_type='application/pdf')
            httpresponse['Content-Disposition'] = 'attachment; filename="{}.pdf"'.format(doc_title)
            table=report_gen(contents, doc_title, httpresponse,httprequest.user.username)

        return httpresponse
    return render(httprequest, "stock_web/listinventory.html", context)

@user_passes_test(is_logged_in, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def stockreport(httprequest, fin, filters, pk, extension):
    submiturl = reverse("stock_web:stockreport",args=[fin, filters, pk,extension])
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Stock Reports")
    header = "Select Reagent to Generate Stock Report For"
    form=StockReportForm
    if pk=="_":
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or "Download" not in httprequest.POST["submit"]:
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST)
                if form.is_valid():
                    queries = []
                    for key, query in [( "rec_range","date_rec__range"), ("open_range","date_op__range"),
                                       ("val_range","val_id__val_date__range"), ("fin_range","date_fin__range"),
                                      ]:
                        val = form.cleaned_data[key]
                        if val:
                            if val[0]!=None:
                                if "range" in query:
                                    val=(val[0].strftime("%Y-%m-%d"), val[1].strftime("%Y-%m-%d"))
                                queries += ["{}={}".format(query, val)]
                    if queries==[]:
                        queries="_"
                    if "pdf" in httprequest.POST["submit"]:
                        return HttpResponseRedirect(reverse("stock_web:stockreport", args=[form.cleaned_data["in_stock"], ";".join(queries), form.cleaned_data["name"].pk,0]))
                    elif "xlsx" in httprequest.POST["submit"]:
                        return HttpResponseRedirect(reverse("stock_web:stockreport", args=[form.cleaned_data["in_stock"], ";".join(queries), form.cleaned_data["name"].pk,1]))
        else:
            form = form(initial = {"in_stock":1})

    else:
        title="{} - Stock Report - Downloaded {}".format(Reagents.objects.get(pk=int(pk)), datetime.datetime.today().date().strftime("%d-%m-%Y"))
        #gets items, with open items first, then sorted by expirey date
        items = Inventory.objects.select_related("supplier","reagent","internal","val","op_user", "team").filter(reagent_id=int(pk),finished__lte=fin).order_by("-is_op","date_exp")
        if filters!="_":
            query = dict([q.split("=") for q in filters.split(";")])
            for key, value in query.items():
                query[key]=value.strip("()").replace("'","").replace(" ","").split(",")
            items = items.filter(**query)
        if len(items)==0:
            messages.success(httprequest, "No inventory items fit the search criteria")
            return HttpResponseRedirect(reverse("stock_web:stockreport", args=["_","_","_","_"]))
        body=[["Supplier Name", "Catalogue Number", "Lot Number", "Stock Number","Team", "Date Received",
               "Expiry Date", "Date Open", "Opened By", "Date Validated", "Validation Run"]]
        if fin=="1":
            body[-1].append("Date Finished")
            body[-1].append("Finished By")
        for item in items:
            body+= [[ item.supplier.name,
                      item.reagent.cat_no,
                      item.lot_no,
                      item.internal.batch_number,
                      item.team.name,
                      item.date_rec.strftime("%d/%m/%Y"),
                      item.date_exp.strftime("%d/%m/%Y"),
                      item.date_op.strftime("%d/%m/%Y") if item.date_op is not None else "",
                      item.op_user.username if item.op_user is not None else "",
                      item.val.val_date.strftime("%d/%m/%Y") if item.val is not None else "",
                      item.val.val_run if item.val is not None else "",
                      ]]
            body[-1].append(item.date_fin.strftime("%d/%m/%Y") if item.date_fin is not None else "")
            body[-1].append(item.fin_user.username if item.fin_user is not None else "")
        if extension=='0':
            httpresponse = HttpResponse(content_type='application/pdf')
            httpresponse['Content-Disposition'] = 'attachment; filename="{}.pdf"'.format(title)
            table=report_gen(body,title,httpresponse,httprequest.user.username)

        if extension=='1':
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            for row in body:
                worksheet.append(row)
            httpresponse = HttpResponse(content=openpyxl.writer.excel.save_virtual_workbook(workbook), content_type='application/ms-excel')
            httpresponse['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(title)
        return httpresponse
    return render(httprequest, "stock_web/reportform.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def invreport(httprequest, team, filters, what, extension):
    submiturl = reverse("stock_web:invreport",args=[type, filters, what,extension])
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Inventory Reports")
    form=InvReportForm
    header = "Select Inventory Report To Generate"
    if what=="_":
        if httprequest.method=="POST":

            if "submit" not in httprequest.POST or "Download" not in httprequest.POST["submit"]:
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST)
                if form.is_valid():
                    queries = []
                    for key, query in [( "rec_range","date_rec__range"), ("open_range","date_op__range"),
                                       ("val_range","val_id__val_date__range"), ("fin_range","date_fin__range"),
                                      ]:
                        val = form.cleaned_data[key]
                        if val:
                            if val[0]!=None:
                                if "range" in query:
                                    val=(val[0].strftime("%Y-%m-%d"), val[1].strftime("%Y-%m-%d"))
                                queries += ["{}={}".format(query, val)]
                    if queries==[]:
                        queries="_"
                    if "pdf" in httprequest.POST["submit"]:
                        return HttpResponseRedirect(reverse("stock_web:invreport", args=[form.cleaned_data["team"].id if form.cleaned_data["team"] is not None else "ALL", ";".join(queries), form.cleaned_data["report"],0]))
                    elif "xlsx" in httprequest.POST["submit"]:
                        return HttpResponseRedirect(reverse("stock_web:invreport", args=[form.cleaned_data["team"].id if form.cleaned_data["team"] is not None else "ALL", ";".join(queries), form.cleaned_data["report"],1]))
        else:
            form = form()
    else:
        if filters!="_":
            query = dict([q.split("=") for q in filters.split(";")])
            for key, value in query.items():
                query[key]=value.strip("()").replace("'","").replace(" ","").split(",")
        elif what=="exp":
            title="Items Expiring Soon Report - Downloaded {}".format(datetime.datetime.today().date().strftime("%d-%m-%Y"))
            items = Inventory.objects.select_related("supplier","reagent","internal","val","op_user").filter(date_exp__lte=datetime.datetime.now()+datetime.timedelta(days=42),finished=False).order_by("reagent_id__name","-is_op","date_exp")
            if team!="ALL":
                items=items.filter(team=team)
            if filters!="_":
                items=items.filter(**query)
            if len(items)==0:
                messages.success(httprequest, "No inventory items fit the search criteria")
                return HttpResponseRedirect(reverse("stock_web:invreport", args=["_","_","_","_"]))
            body=[["Reagent", "Catalogue Number", "Current Volume", "Team", "Supplier", "Lot Number", "Stock Number", "Received",
                   "Expiry", "Opened", "Opened By", "Date Validated", "Validation Run"]]
            for item in items:
                body+= [[item.reagent.name,
                          item.reagent.cat_no,
                          "{}µl".format(item.current_vol) if item.current_vol is not None else "N/A",
                          item.team.name,
                          item.supplier.name,
                          item.lot_no,
                          item.internal.batch_number,
                          item.date_rec.strftime("%d/%m/%Y"),
                          item.date_exp.strftime("%d/%m/%Y"),
                          item.date_op.strftime("%d/%m/%Y") if item.date_op is not None else "",
                          item.op_user.username if item.op_user is not None else "",
                          item.val.val_date.strftime("%d/%m/%Y") if item.val is not None else "",
                          item.val.val_run if item.val is not None else "",
                          ]]
        elif what=="minstock":
            title="Items Below Their Minimum Stock Levels - Downloaded {}".format(datetime.datetime.today().date().strftime("%d-%m-%Y"))
            items=Reagents.objects.filter(count_no__lt=F("min_count")).order_by("name")
            if team!="ALL":
                items=items.filter(team_def=team)
            if len(items)==0:
                messages.success(httprequest, "No inventory items fit the search criteria")
                return HttpResponseRedirect(reverse("stock_web:invreport", args=["_","_","_","_"]))
            body=[["Reagent", "Catalogue Number", "Default Team", "Default Supplier", "Number In Stock", "Minimum Stock Level"]]
            for item in items:
                body+= [[item.name,
                        item.cat_no,
                        item.team_def.name if item.team_def.name is not None else "",
                        item.supplier_def.name if item.supplier_def is not None else "",
                         "{}µl".format(item.count_no) if item.track_vol==True else item.count_no,
                         "{}µl".format(item.min_count) if item.track_vol==True else item.min_count]]
        if extension=='0':
            httpresponse = HttpResponse(content_type='application/pdf')
            httpresponse['Content-Disposition'] = 'attachment; filename="{}.pdf"'.format(title)
            table=report_gen(body,title,httpresponse,httprequest.user.username)

        if extension=='1':
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            for row in body:
                worksheet.append(row)
            httpresponse = HttpResponse(content=openpyxl.writer.excel.save_virtual_workbook(workbook), content_type='application/ms-excel')
            httpresponse['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(title)
        return httpresponse
    return render(httprequest, "stock_web/reportform.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})

def _item_context(httprequest, item, undo):
    title = ["Reagent - {}".format(item.reagent.name),
             "Supplier - {}".format(item.supplier.name),
             "Catalogue Number - {}".format(item.reagent.cat_no) if item.sol is None else "",
             "Team - {}".format(item.team.name),
             "Lot Number - {}".format(item.lot_no) if item.lot_no !="N/A" else "",
             "Stock Number - {}".format(item.internal.batch_number)]
    title_url=["","","","","",""]
    if undo=="undo":
        title[0:0]=["***WARNING - ONLY TO BE USED TO CORRECT DATA ENTRY ERRORS. IT MAY NOT BE POSSIBLE TO UNDO CHANGES MADE HERE***"]
        title_url.append("")
    if item.po is not None:
        title.append("Purchase Order Number - {}".format(item.po))
        title_url.append("")
    if item.accept_reason is not None:
        title.append("Reason for Acceptance - {}".format(item.accept_reason))
        title_url.append("")
    if item.witness is not None:
        title.append("Witnessed By - {}".format(item.witness))
        title_url.append("")
    if item.sol is not None and undo!="undo":
        for comp in item.sol.list_comp():
            if comp.val_id is not None:
                title.append(comp)
            else:
                title.append(str(comp)+" - NOT VALIDATED")
            title_url.append(reverse("stock_web:item",args=[comp.id]))
    if item.val is None:
        title.append("****ITEM NOT VALIDATED****")
        title_url.append("")
    title=zip(title,title_url)
    if item.sol is not None:
        headings = ["Date Created", "Created By", "Condition Received", "Expiry Date"]
    else:
        headings = ["Date Received", "Received By", "Condition Received", "Expiry Date"]
    values = [item.date_rec, item.rec_user.username, CONDITIONS[item.cond_rec], item.date_exp]
    urls = ["", "", "", ""]
    SKIP=False
    if item.date_op is not None:
        headings+=["Date Opened", "Opened By"]
        values+=[item.date_op,item.op_user]
        urls+=["",""]
    if item.val_id is not None:
        headings+=["Date Validated", "Validation Run", "Validation User"]
        values+=[item.val.val_date, item.val.val_run, item.val.val_user]
        urls+=["","",""]
    if ((item.date_op is None) and (item.finished==False)):
        headings+=["Action"]
        if undo=="undo":
            values+=["Delete Item"]
            urls+=[reverse("stock_web:undoitem",args=["delete",item.id])]
        else:
            values+=["Open Item"]
            urls+=[reverse("stock_web:openitem",args=[item.id])]
    if ((item.date_op is not None) and (item.val_id is None) and (item.finished==False)):
        if undo=="undo":
            headings+=["Action"]
            values+=["Un-open Item"]
            urls+=[reverse("stock_web:undoitem",args=["unopen",item.id])]
        elif item.sol_id is None and (httprequest.user.groups.filter(name="Admin").exists() or httprequest.user.is_staff):
            headings+=["Action"]
            values+=["Validate Item"]
            urls+=[reverse("stock_web:valitem",args=[item.id])]
    elif ((item.date_op is not None) and (item.val_id is not None) and (item.finished==False) and (undo=="undo")):
        headings+=["Action"]
        values+=["Un-open Item"]
        urls+=[reverse("stock_web:undoitem",args=["unopen",item.id])]
    if item.val_id is not None and item.finished==False:
        if undo=="undo":
            headings+=["Action"]
            values+=["Un-Validate Item"]
            urls+=[reverse("stock_web:undoitem",args=["unval",item.id])]
        elif item.is_op==True:
            headings+=["Action"]
            values+=["Finish/Discard Item"]
            urls+=[reverse("stock_web:finishitem",args=[item.id])]
            SKIP=True

    if ((item.finished==False) and (undo!="undo")  and (SKIP==False)):
        headings+=["Action"]
        if item.sol_id is not None:
            values+=["Finish Item"]
        else:
            values+=["Discard Item"]
        urls+=[reverse("stock_web:finishitem",args=[item.id])]
    if item.finished==True:
        if item.is_op==True:
            headings+=["Date Finished", "Finished by"]
        else:
            headings+=["Date Discarded", "Discared by"]
        values+=[item.date_fin, item.fin_user]
        urls+=["",""]
        if undo=="undo":
            headings+=["Action"]
            values+=["Re-Open Item"]
            urls+=[reverse("stock_web:undoitem",args=["reopen",item.id])]
    body = [(zip(values,urls, urls),False)]
    context = {"header":title,"headings":headings, "body":body, "toolbar":_toolbar(httprequest), "cyto":False}
    if ((item.finished==True) and (item.fin_text is not None)):
        context.update({"newinformation":item.fin_text})
    return context

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def _vol_context(httprequest, item, undo):
    stripe=False
    title = ["Reagent - {}".format(item.reagent.name),
             "Supplier - {}".format(item.supplier.name),
             "Catalogue Number - {}".format(item.reagent.cat_no) if item.sol is None else "",
             "Team - {}".format(item.team.name),
             "Purchase Order Number - {}".format(item.po),
             "Lot Number - {}".format(item.lot_no) if item.lot_no!="N/A" else "",
             "Stock Number - {}".format(item.internal.batch_number),
             "Volume Received - {}µl".format(item.vol_rec) if item.sol is None else "Volume Made Up - {}µl".format(item.vol_rec),
             "Current Volume - {}µl".format(item.current_vol if item.current_vol is not None else 0)]
    title_url=["","","","","","","","",""]
    skip=False
    if undo=="undo":
        title[0:0]=["***WARNING - ONLY TO BE USED TO CORRECT DATA ENTRY ERRORS. IT MAY NOT BE POSSIBLE TO UNDO CHANGES MADE HERE***"]
        title_url.append("")
    if item.accept_reason is not None:
        title.append("Reason for Acceptance - {}".format(item.accept_reason))
        title_url.append("")
    if item.witness is not None:
        title.append("Witnessed By - {}".format(item.witness))
        title_url.append("")
    if item.sol is not None:
        for comp in item.sol.list_comp():
            if comp.val_id is not None:
                title.append(comp)
            else:
                title.append(str(comp)+" - NOT VALIDATED")
            title_url.append(reverse("stock_web:item",args=[comp.id]))
    if item.val is None:
        title.append("****ITEM NOT VALIDATED****")
        title_url.append("")
    title=zip(title,title_url)
    if item.sol is not None:
        headings = ["Date Created", "Created By", "Condition Received", "Expiry Date"]
    else:
        headings = ["Date Received", "Received By", "Condition Received", "Expiry Date"]
    values = [item.date_rec, item.rec_user.username, CONDITIONS[item.cond_rec], item.date_exp]
    urls = ["", "", "", ""]
    SKIP=False
    if item.date_op is not None:
        headings+=["Date Opened", "Opened By"]
        values+=[item.date_op,item.op_user]
        urls+=["",""]
    if item.val_id is not None:
        headings+=["Date Validated", "Validation Run", "Validation User"]
        values+=[item.val.val_date, item.val.val_run, item.val.val_user]
        urls+=["","",""]
    if ((item.date_op is None) and (item.finished==False)):
        headings+=["Action"]
        if undo=="undo":
            values+=["Delete Item"]
            urls+=[reverse("stock_web:undoitem",args=["delete",item.id])]
        else:
            values+=["Open Item"]
            urls+=[reverse("stock_web:openitem",args=[item.id])]
    if ((item.is_op==True) and (item.finished==False) and (undo!="undo")):
        headings+=["Action"]
        values+=["Use Amount"]
        urls+=[reverse("stock_web:useitem",args=[item.id])]
    if ((item.date_op is not None) and (item.val_id is None) and (item.finished==False)):
        if undo=="undo":
            headings+=["Action"]
            values+=["Un-open Item"]
            urls+=[reverse("stock_web:undoitem",args=["unopen",item.id])]
        elif item.sol_id is None and item.last_usage is not None and (httprequest.user.groups.filter(name="Admin").exists() or httprequest.user.is_staff):
            headings+=["Action"]
            values+=["Validate Item"]
            urls+=[reverse("stock_web:valitem",args=[item.id])]
    elif ((item.date_op is not None) and (item.val_id is not None) and (item.finished==False) and (undo=="undo")):
        headings+=["Action"]
        values+=["Un-open Item"]
        urls+=[reverse("stock_web:undoitem",args=["unopen",item.id])]
    if item.val_id is not None and item.finished==False:
        if undo=="undo":
            headings+=["Action"]
            values+=["Un-Validate Item"]
            urls+=[reverse("stock_web:undoitem",args=["unval",item.id])]
        elif item.is_op==True:
            headings+=["Action"]
            values+=["Finish/Discard Item"]
            urls+=[reverse("stock_web:finishitem",args=[item.id])]
            SKIP=True
    if ((item.finished==False) and (undo!="undo")  and (SKIP==False)):
        headings+=["Action"]
        values+=["Discard Item"]
        urls+=[reverse("stock_web:finishitem",args=[item.id])]
    if item.finished==True:
        if item.is_op==True:
            headings+=["Date Finished", "Finished by"]
        else:
            headings+=["Date Discarded", "Discared by"]
        values+=[item.date_fin, item.fin_user]
        urls+=["",""]
    body = [(zip(values,urls, urls),stripe)]
    context = {"header":title,"headings":headings, "body":body, "toolbar":_toolbar(httprequest)}
    if ((item.finished==True) and (item.fin_text is not None)):
        context.update({"newinformation":item.fin_text})
    if item.last_usage is not None:
        cyto_headings=["Volume at Start", "Volume at End", "Volume Used", "Date", "User"]
        if undo=="undo":
            cyto_headings+=["Action"]
        uses=VolUsage.objects.filter(item=item.pk)
        uses=sorted(uses, key = lambda use:use.date)
        cyto_body=[]
        for use in uses:
            values=["{}µl".format(use.start),
                    "{}µl".format(use.end),
                    "{}µl".format(use.used),
                    use.date,
                    use.user]
            urls=["","","","",""]
            if use.sol is not None:
                urls=[reverse("stock_web:item",args=[Inventory.objects.get(sol=use.sol).pk])]*5
            if undo=="undo":
                if use==item.last_usage:
                    if item.finished==True:
                        values+=["EDIT (RE-OPEN)"]
                    else:
                        values+=["EDIT"]
                    urls+=[reverse("stock_web:undoitem",args=["unuse",item.id])]
                else:
                    values+=[""]
                    urls+=[""]
            cyto_body.append((zip(values,urls),stripe, True if use.used<0 else False))
            stripe=not(stripe)
        context.update({"cyto":True,
                        "cyto_headings":cyto_headings,
                        "cyto_body":cyto_body})
    return context

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def useitem(httprequest,pk):
    item=Inventory.objects.get(pk=int(pk))
    if item.is_op==False:
        return HttpResponseRedirect(reverse("stock_web:item",args=[pk]))
    uses=VolUsage.objects.filter(item=item)
    if item.is_op==True and item.val is None and len(uses)>0 and item.sol is None:
        messages.success(httprequest, "WARNING - ITEM IS NOT VALIDATED")
    form=UseItemForm
    header=["Enter Volume used for {} - Current Volume is {}µl".format(item, item.current_vol)]
    if httprequest.method=="POST":
        form = form(httprequest.POST, instance=item)
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:

            if form.is_valid():
                Inventory.take_out(form.cleaned_data["vol_used"], int(pk), httprequest.user, form.cleaned_data["date_used"])
                item.refresh_from_db()
                message=[]
                if item.reagent.count_no<item.reagent.min_count:
                    if item.sol is not None:
                        make="made"
                    else:
                        make="ordered"
                    message+=["Current stock level for {} is {}µl. Minimum quantity is {}µl. Check if more needs to be {}".format(item.reagent.name,item.reagent.count_no,
                                                                                                                          item.reagent.min_count, make)]


                    if EMAIL==True:
                        subject="{} - Stock Level is below minimum level".format(item.reagent.name)
                        text="<p>Item {} has a stock level of {}µl.<br><br>".format(item.reagent.name,item.reagent.count_no)
                        text+="Minimum Stock level for this item is {}µl.<br><br>".format(item.reagent.min_count)
                        for user in User.objects.filter(is_staff=True, is_active=True):
                            if user.email!="":
                                try:
                                    send(subject,text, user.email)
                                except Exception as e:
                                    print(e)
                if int(item.current_vol)==0:
                    message+=["THIS TUBE IS EMPTY, PLEASE DISCARD IT!"]
                if form.cleaned_data["date_used"]>=item.date_exp:
                    message+=["WARNING - ITEM USED AFTER EXPIRY DATE"]
                if message!=[]:
                    messages.success(httprequest," \n".join(message))

                return HttpResponseRedirect(reverse("stock_web:item",args=[pk]))
    else:
        form=form(instance=item,initial = {"date_used":datetime.datetime.now()})
    submiturl = reverse("stock_web:useitem",args=[pk])
    cancelurl = reverse("stock_web:item",args=[pk])
    return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": _toolbar(httprequest), "submiturl": submiturl, "cancelurl": cancelurl})


@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def openitem(httprequest, pk):
    item=Inventory.objects.get(pk=int(pk))
    form=OpenItemForm
    header=["Opening item {}".format(item)]
    header+=["Date Received: {}".format(item.date_rec)]
    if httprequest.method=="POST":
        form = form(httprequest.POST, instance=item)
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            if form.is_valid():
                Inventory.open(form.cleaned_data, pk, httprequest.user)
                item.refresh_from_db()
                messages_to_show=[]
                if item.reagent.track_vol==False:
                    if item.reagent.count_no<item.reagent.min_count:
                        if item.sol is not None:
                            make="made"
                        else:
                            make="ordered"
                        messages_to_show+=["Current stock level for {} is {}. Minimum quantity is {}. Check if more needs to be {}".format(item.reagent.name,
                                                                                                                                              item.reagent.count_no,
                                                                                                                                              item.reagent.min_count,
                                                                                                                                              make)]
                        if EMAIL==True:
                            subject="{} - Stock Level is below minimum level".format(item.reagent.name)
                            text="<p>Item {} has a stock level of {}.<br><br>".format(item.reagent.name,item.reagent.count_no)
                            text+="Minimum Stock level for this item is {}.<br><br>".format(item.reagent.min_count)
                            for user in User.objects.filter(is_staff=True, is_active=True):
                                if user.email!="":
                                    try:
                                        send(subject,text, user.email)
                                    except Exception as e:
                                        print(e)
                #Shows a warning if the item is opened after it's expiry date
                if form.cleaned_data["date_op"]>=item.date_exp:
                    messages_to_show+=["WARNING - ITEM OPEN AFTER EXPIRY DATE"]
                if messages_to_show!=[]:
                    messages.success(httprequest,"\n".join(messages_to_show))
                return HttpResponseRedirect(reverse("stock_web:item",args=[pk]))
    else:
        if item.is_op==True:
            return HttpResponseRedirect(reverse("stock_web:item",args=[pk]))
        form=form(instance=item,initial = {"date_op":datetime.datetime.now()})
    submiturl = reverse("stock_web:openitem",args=[pk])
    cancelurl = reverse("stock_web:item",args=[pk])
    return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": _toolbar(httprequest), "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def valitem(httprequest,pk):
    item=Inventory.objects.get(pk=int(pk))
    form=ValItemForm
    if Inventory.objects.get(pk=int(pk)).is_op==False:
        return HttpResponseRedirect(reverse("stock_web:item",args=[pk]))
    header=["Validating item {}".format(item)]
    header+=["Date Open: {}".format(item.date_op)]
    if httprequest.method=="POST":
        form = form(httprequest.POST, instance=item)
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            if form.is_valid():
                validated=Inventory.validate(form.cleaned_data, Inventory.objects.get(pk=int(pk)).reagent, Inventory.objects.get(pk=int(pk)).lot_no, httprequest.user)
                val_list=[x.internal.batch_number for x in validated]
                messages.success(httprequest, "Validated: " + " \n".join(val_list))
                return HttpResponseRedirect(reverse("stock_web:item",args=[pk]))
    else:
        if Inventory.objects.get(pk=int(pk)).val is not None:
            return HttpResponseRedirect(reverse("stock_web:item",args=[pk]))
        form=form(instance=item,initial = {"val_date":datetime.datetime.now()})
    submiturl = reverse("stock_web:valitem",args=[pk])
    cancelurl = reverse("stock_web:item",args=[pk])
    return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": _toolbar(httprequest), "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def finishitem(httprequest, pk):
    item=Inventory.objects.get(pk=int(pk))
    form=FinishItemForm
    header=["Finishing item {}".format(item if item.sol_id is None else ", ".join(str(item).split(',')[::2]))]
    header+=["Date Open: {}".format(item.date_op if item.is_op==True else "NOT OPEN")]
    header+=["Date Validated: {}".format(item.val.val_date if item.val is not None else "NOT REQUIRED" if item.sol_id is not None else "NOT VALIDATED")]
    if httprequest.method=="POST":
        form = form(httprequest.POST, instance=item)
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            if form.is_valid():

                Inventory.finish(form.cleaned_data, pk, httprequest.user)
                if item.reagent.track_vol==True or item.is_op==False:
                    if item.reagent.count_no<item.reagent.min_count:
                        if item.sol is not None:
                            make="made"
                        else:
                            make="ordered"
                        messages.success(httprequest, "Current stock level for {0} is {1}{2}. \nMinimum quantity is {3}{2}. \nCheck if more needs to be {4}".format(item.reagent.name,
                                                                                                                                              item.reagent.count_no,
                                                                                                                                              "µl" if item.reagent.track_vol else "",
                                                                                                                                              item.reagent.min_count,
                                                                                                                                              make))
                        if EMAIL==True:
                            subject="{} - Stock Level is below minimum level".format(item.reagent.name)
                            text="<p>Item {} has a stock level of {}.<br><br>".format(item.reagent.name,item.reagent.count_no)
                            text+="\n\nMinimum Stock level for this item is {}.<br><br>".format(item.reagent.min_count)
                            for user in User.objects.filter(is_staff=True, is_active=True):
                                if user.email!="":
                                    try:
                                        send(subject,text, user.email)
                                    except Exception as e:
                                        print(e)

                if item.val_id is None and item.is_op==True and item.sol is None:
                    if EMAIL==True:
                        subject="{} - Discarded without validation".format(item.reagent.name)
                        text="<p>Item {} ({}) has been discarded by {} without having validation data.<br><br>".format(item.reagent.name,item.internal.batch_number,httprequest.user.username)
                        text+="\n\nThe reason they entered was: '{}'<br><br>".format(form.cleaned_data["fin_text"] if form.cleaned_data["fin_text"]!=None else "NOT ENTERED")
                        for user in User.objects.filter(is_staff=True, is_active=True):
                            if user.email!="":
                                try:
                                    send(subject,text, user.email)
                                except Exception as e:
                                    print(e)
                return HttpResponseRedirect(reverse("stock_web:item",args=[pk]))
    else:
        if Inventory.objects.get(pk=int(pk)).finished==True:
            return HttpResponseRedirect(reverse("stock_web:item",args=[pk]))
        if item.is_op==False:
            messages.success(httprequest,"WARNING - ITEM HAS NOT BEEN OPENED")
        if item.val_id is None and item.is_op==True and item.sol is None:
            messages.success(httprequest,"WARNING - THIS ITEM HAS NOT BEEN VALIDATED")
        form=form(instance=item,initial = {"date_fin":datetime.datetime.now()})
    submiturl = reverse("stock_web:finishitem",args=[pk])
    cancelurl = reverse("stock_web:item",args=[pk])
    return render(httprequest, "stock_web/form.html", {"header": header,"form": form, "toolbar": _toolbar(httprequest), "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def item(httprequest, pk):
    item = Inventory.objects.select_related("supplier","reagent","internal","val").get(pk=int(pk))
    if item.reagent.track_vol==False:
        return render(httprequest, "stock_web/list_item.html", _item_context(httprequest, item, "_"))
    else:
        return render(httprequest, "stock_web/list_item.html", _vol_context(httprequest, item, "_"))

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def recipes(httprequest):
    title = "List of Recipes"
    headings = ["Recipe Name", "Number of Components", "Shelf Life (Months)", "Active", "Witness Required?"]
    items=Recipe.objects.all().order_by("name")
    body=[]

    for item in items:
        values = [item.name,
                  item.length,
                  item.shelf_life,
                  "YES" if item.reagent.is_active else "NO",
                  "YES" if item.witness_req else "NO",]
        urls=[reverse("stock_web:recipe",args=[item.id]),
              "",
              "",
              "",
              "",
              ]
        body.append((zip(values,urls), False))

    context = {"header":title,"headings":headings, "body":body, "toolbar":_toolbar(httprequest, active="Recipes")}
    return render(httprequest, "stock_web/list.html", context)

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def recipe(httprequest, pk):
    item = Recipe.objects.get(pk=int(pk))
    title = "Components for {}".format(item.name)
    headings = ["Reagent", "Amount in stock"]
    body=[]
    for i in range(1, item.length()+1):
        values = [eval('item.comp{}.name'.format(i)),
                  "{}µl".format(eval('item.comp{}.count_no'.format(i))) if eval('item.comp{}.track_vol'.format(i))==True else eval('item.comp{}.count_no'.format(i)),
                  ]
        urls= ["",
               ""]
        body.append((zip(values,urls),False))
    context = {"header":title,"headings":headings, "body":body, "toolbar":_toolbar(httprequest, active="Recipe")}
    return render(httprequest, "stock_web/list.html", context)

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def newinv(httprequest, pk):
    if pk=="_":
        title="Select Reagent to Book-in"
        template="stock_web/invform.html"
        form=NewInvForm1
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or httprequest.POST["submit"] != "Book-in":
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST)
                if form.is_valid():
                    if form.cleaned_data["reagent"].recipe is None:
                        return HttpResponseRedirect(reverse("stock_web:newinv", args=[form.cleaned_data["reagent"].pk]))
                    elif form.cleaned_data["reagent"].recipe is not None:
                        return HttpResponseRedirect(reverse("stock_web:createnewsol", args=[form.cleaned_data["reagent"].recipe.pk]))
        else:
            form = form()
    else:
        item=Reagents.objects.get(pk=int(pk))
        if item.recipe is not None:
            return HttpResponseRedirect(reverse("stock_web:createnewsol", args=[item.recipe_id]))
        title=["Enter Delivery Details - {} {}".format(item, "- " + item.cat_no if item.cat_no is not None else "")]
        template="stock_web/newinvform.html"
        if item.track_vol==False:
            form=NewInvForm
        elif item.track_vol==True:
            form=NewProbeForm
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST,instance=item)
                if form.is_valid():
                    ids=Inventory.create(form.cleaned_data, httprequest.user)
                    message=[]
                    if item.track_vol==False:
                        quant=form.cleaned_data["reagent"].count_no+int(form.data["num_rec"])
                    elif item.track_vol==True:
                        quant=form.cleaned_data["reagent"].count_no+int(form.data["vol_rec"])
                    if quant<form.cleaned_data["reagent"].min_count:
                        if form.cleaned_data["reagent"].recipe is not None:
                            make="made"
                        else:
                            make="ordered"
                        message+=["Current stock level for {} is {}. Minimum quantity is {}. Check if more needs to be {}".format(form.cleaned_data["reagent"].name,
                                                                                                                                       quant,
                                                                                                                                       form.cleaned_data["reagent"].min_count,
                                                                                                                                       make)]
                    if item.recipe is None:
                        items=Inventory.objects.filter(reagent=form.cleaned_data["reagent"].id, lot_no=form.cleaned_data["lot_no"],val_id__gte=0)
                        if len(items)>0:
                             message+=["THIS ITEM IS VALIDATED. RUN {}".format(items[0].val.val_run)]
                        else:
                             message+=["THIS ITEM IS NOT VALIDATED"]
                        if item.track_vol==False:
                            messages.error(httprequest, "{}x {} added".format(form.data["num_rec"],
                                                                              form.cleaned_data["reagent"]))
                        elif item.track_vol==True:
                            messages.error(httprequest, "1x {}µl of {} added".format(form.data["vol_rec"],
                                                                              form.cleaned_data["reagent"]))
                    if form.cleaned_data["date_exp"]<(form.cleaned_data["date_rec"] + relativedelta(months=+6)):
                        message+=["ITEM EXPIRES WITHIN 6 MONTHS"]
                    if message!=[]:
                        messages.success(httprequest," \n".join(message))
                    messages.error(httprequest, "STOCK NUMBERS:")
                    for ID in ids:
                        messages.info(httprequest, ID)
                    return HttpResponseRedirect(reverse("stock_web:newinv",args=["_"]))
        else:
            form = form(initial = {"supplier":item.supplier_def,
                                   "team":item.team_def,
                                   "reagent":item})
    submiturl = reverse("stock_web:newinv",args=[pk])
    cancelurl = reverse("stock_web:listinv")
    if httprequest.user.is_staff:
        active="new"
    else:
        active="New Inventory Item"
    return render(httprequest, template, {"header":title, "form": form, "toolbar": _toolbar(httprequest, active=active), "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def createnewsol(httprequest, pk):
    recipe = Recipe.objects.get(pk=int(pk))
    title="Select Items for Recipe: {}".format(recipe)
    if recipe.witness_req==True:
        form=WitnessForm
    else:
        form=TeamOnlyForm
    if httprequest.method == "POST":
        if "submit" in httprequest.POST and httprequest.POST["submit"] == "save":
            form = form(httprequest.POST)
            Reagents=[Inventory.objects.get(pk=int(x)).reagent for x in httprequest.POST.getlist("requests")]
            if len(Reagents)>10:
                messages.success(httprequest, "More Than 10 Items Selected. Max Component Limit is 10 Items")
                return HttpResponseRedirect(reverse("stock_web:createnewsol",args=[pk]))
            Reagents_set=set(Reagents)
            vols_used={}
            vol_made=""
            potentials=recipe.liststock()
            potentials.sort(key=attrgetter("is_op"),reverse=True)
            comp_vol=any(p.current_vol is not None for p in potentials)
            witness=None
            team=form.data["team"]
            try:
                witness=User.objects.get(pk=int(form.data["name"]))
                if witness==httprequest.user:
                    messages.success(httprequest, "YOU MAY NOT USE YOURSELF AS A WITNESS")
                    return HttpResponseRedirect(reverse("stock_web:createnewsol",args=[pk]))
            except (ValueError, KeyError):
                witness=None
            if recipe.track_vol==True:
                vol_made=httprequest.POST.getlist("total_volume")[0]
                if httprequest.POST.getlist("total_volume")==[""]:
                    messages.success(httprequest, "Total Volume Made Not Entered")
                    return HttpResponseRedirect(reverse("stock_web:createnewsol",args=[pk]))
            if comp_vol==True:
                if all(v=="" for v in httprequest.POST.getlist("volume")):
                    messages.success(httprequest, "No Volumes Entered")
                    return HttpResponseRedirect(reverse("stock_web:createnewsol",args=[pk]))
                #if the first item isn't a cyto, zip gives the first cyto (item #2) to item #1
                #work around was to give everything that's not cyto a hidden 0, but then those all of those aren't checked so ERRORS
                #find fix for allowing intended 0+ticked to be counted but not giving "volumes entered doesn't match tick boxes" error...?
                vols=zip(potentials,httprequest.POST.getlist("volume"))

                for vol in vols:
                    #skips volumes with "a" as a is the value given to hidden volumes so the zip works properly
                    if vol[1]=="a":
                        continue
                    if vol[1]!="":
                        if str(vol[0].pk) not in httprequest.POST.getlist("requests"):
                            messages.success(httprequest, "Selected Checkmarks and Volume Used boxes do not match")
                            return HttpResponseRedirect(reverse("stock_web:createnewsol",args=[pk]))
                        elif vol[1]!="0":
                            vols_used[str(vol[0].pk)]=vol[1]
                for req in httprequest.POST.getlist("requests"):
                    if req not in vols_used.keys() and Inventory.objects.get(pk=int(req)).current_vol is not None:
                        messages.success(httprequest, "Selected Checkmarks and Volume Used boxes do not match")
                        return HttpResponseRedirect(reverse("stock_web:createnewsol",args=[pk]))

                errors=[]
                sum_vol=0
                for item, vol in vols_used.items():
                    invitem=Inventory.objects.get(pk=int(item))
                    sum_vol+=int(vol)
                    if int(invitem.current_vol)-int(vol)<0:
                        errors+=["Reagent {} only has {}µl in the tube. Cannot take {}µl".format(invitem.reagent.name, invitem.current_vol, vol)]
                if errors!=[]:
                    messages.success(httprequest, " ".join(errors))
                    return HttpResponseRedirect(reverse("stock_web:createnewsol",args=[pk]))
                if recipe.track_vol==True:
                    if int(vol_made)<sum_vol:
                        messages.success(httprequest, "Total Volume of Reagents Used is {}µl. Total Volume made must be at least this volume".format(sum_vol))
                        return HttpResponseRedirect(reverse("stock_web:createnewsol",args=[pk]))



            if len(Reagents_set)!=recipe.length():
                if len(Reagents_set)==1:
                    grammar="item was"
                else:
                    grammar="items were"
                messages.success(httprequest, "Only {} of the required items {} selected. This recipe requires {} items".format(len(Reagents_set),
                                                                                                          grammar,
                                                                                                          recipe.length()))
                return HttpResponseRedirect(reverse("stock_web:createnewsol",args=[pk]))
            messages_to_show=[]
            for item in [Inventory.objects.get(pk=int(x)) for x in httprequest.POST.getlist("requests")]:
                if item.is_op==False:
                    messages_to_show+=["Reagent {} was not previously open. It has now been marked as open on its date received".format(item)]
            sol, changed, EXP_DATE=Solutions.create(recipe, [int(x) for x in httprequest.POST.getlist("requests") if x.isdigit()], vols_used, vol_made, httprequest.user, witness, team)
            if changed==True:
                messages_to_show+=["A component has an expiry date earlier than the shelf life of this solution. New expiry date is {}".format(EXP_DATE.strftime("%d/%m/%Y"))]
            if messages_to_show!=[]:
                messages.success(httprequest,"\n".join(messages_to_show))
            pk=Inventory.objects.get(internal__batch_number=sol[0]).pk
            return HttpResponseRedirect(reverse("stock_web:item", args=[pk]))
        else:
            HttpResponseRedirect(reverse("stock_web:listinv"))
    else:
        form = form(initial = {"team":recipe.reagent.team_def})
        values=[]
        inv_ids = []
        checked = []
        VOL=[]
        potentials=recipe.liststock()
        potentials.sort(key=attrgetter("is_op"),reverse=True)
        comp_vol=any(p.current_vol is not None for p in potentials)
        if comp_vol==False:
            vol=False
            headings = ["Reagent Name", "Supplier", "Expiry Date", "Stock Number", "Lot Number", "Date Open", "Validation Run", "Select", ""]
        elif comp_vol==True:
            vol=True
            headings = ["Reagent Name", "Supplier", "Expiry Date", "Stock Number", "Lot Number", "Date Open", "Current Volume", "Validation Run", "Select", "Volume used (µl)", ""]
        for p in potentials:
            #temp used so that can make array for that item, then insert if it's a cyto reagent
            temp=[p.reagent.name,
                  p.supplier.name,
                  p.date_exp,
                  p.internal.batch_number,
                  p.lot_no,
                  p.date_op if p.date_op is not None else "NOT OPEN",
                  p.val.val_run if p.val is not None else ""]
            if vol==True:
                temp.insert(-1, "{}µl".format(p.current_vol) if p.current_vol is not None else "N/A")
            values.append(temp)
            inv_ids.append(p.id)
            checked.append("")
            VOL.append(p.current_vol is not None)
        context = { "headings": headings,
                    "body": zip(values, inv_ids, checked, VOL),
                    "url": reverse("stock_web:createnewsol", args=[pk]),
                    "toolbar": _toolbar(httprequest),
                    "total":recipe.track_vol,
                    "identifier":title,
                    "form":form,
                    "cancelurl": reverse("stock_web:newinv",args=["_"]),
                  }
        return render(httprequest, "stock_web/populatesol.html", context)


@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def newreagent(httprequest):
    form=NewReagentForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():
                Reagents.create(form.cleaned_data)
                messages.info(httprequest, "{} Added".format(form.cleaned_data["name"]))
                return HttpResponseRedirect(reverse("stock_web:newreagent"))
    else:
        form = form()
    submiturl = reverse("stock_web:newreagent")
    cancelurl = reverse("stock_web:listinv")
    return render(httprequest, "stock_web/form.html", {"header":["New Reagent Input"], "form": form, "toolbar": _toolbar(httprequest, active="new"), "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def newsup(httprequest):
    form=NewSupForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():
                Suppliers.create(form.cleaned_data["name"])
                messages.info(httprequest, "{} Added".format(form.cleaned_data["name"]))
                return HttpResponseRedirect(reverse("stock_web:newsup"))
    else:
        form = form()
    submiturl = reverse("stock_web:newsup")
    cancelurl = reverse("stock_web:listinv")
    return render(httprequest, "stock_web/form.html", {"header":["New Supplier Input"], "form": form, "toolbar": _toolbar(httprequest, active="new"), "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def newteam(httprequest):
    form=NewTeamForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():
                Teams.create(form.cleaned_data["name"])
                messages.info(httprequest, "{} Added".format(form.cleaned_data["name"]))
                return HttpResponseRedirect(reverse("stock_web:newteam"))
    else:
        form = form()
    submiturl = reverse("stock_web:newteam")
    cancelurl = reverse("stock_web:listinv")
    return render(httprequest, "stock_web/form.html", {"header":["New Team Input"], "form": form, "toolbar": _toolbar(httprequest, active="new"), "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def newrecipe(httprequest):
    form=NewRecipeForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():

                Recipe.create(form.cleaned_data)
                messages.info(httprequest, "{} Added".format(form.cleaned_data["name"]))
                return HttpResponseRedirect(reverse("stock_web:newrecipe"))
    else:
        form = form()
    submiturl = reverse("stock_web:newrecipe")
    cancelurl = reverse("stock_web:listinv")
    return render(httprequest, "stock_web/form.html", {"header":["New Recipe Input"], "form": form, "toolbar": _toolbar(httprequest, active="new"), "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def activsup(httprequest):
    header = ["Select Supplier To (De)Activate - THIS WILL NOT AFFECT EXISTING ITEMS"]
    form=EditSupForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():
                if form.cleaned_data["name"].is_active==True:
                   form.cleaned_data["name"].is_active=False
                   message="Supplier {} Has Been Deactivated".format(form.cleaned_data["name"].name)
                else:
                    form.cleaned_data["name"].is_active=True
                    message="Supplier {} Has Been Reactivated".format(form.cleaned_data["name"].name)
                form.cleaned_data["name"].save()
                messages.success(httprequest, message)
                return HttpResponseRedirect(reverse("stock_web:activsup"))
    else:
        form = form()

    submiturl = reverse("stock_web:activsup")
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Edit Data")

    return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl, "active":"admin"})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def activteam(httprequest):
    header = ["Select Team To (De)Activate - THIS WILL NOT AFFECT EXISTING ITEMS"]
    form=EditTeamForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():
                if form.cleaned_data["name"].is_active==True:
                   form.cleaned_data["name"].is_active=False
                   message="Team {} Has Been Deactivated".format(form.cleaned_data["name"].name)
                else:
                    form.cleaned_data["name"].is_active=True
                    message="Team {} Has Been Reactivated".format(form.cleaned_data["name"].name)
                form.cleaned_data["name"].save()
                messages.success(httprequest, message)
                return HttpResponseRedirect(reverse("stock_web:activteam"))
    else:
        form = form()

    submiturl = reverse("stock_web:activteam")
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Edit Data")

    return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl, "active":"admin"})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def activreag(httprequest):
    header = ["Select An Item To (De)Activate - THIS WILL NOT AFFECT EXISTING ITEMS"]
    form=EditReagForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():
                if form.cleaned_data["name"].recipe is None:
                    word="Reagent"
                else:
                    word="Recipe"
                if form.cleaned_data["name"].is_active==True:
                   form.cleaned_data["name"].is_active=False
                   form.cleaned_data["name"].supplier_def_id=None
                   message="{} {} Has Been Deactivated".format(word, form.cleaned_data["name"].name)
                else:
                    form.cleaned_data["name"].is_active=True
                    if form.cleaned_data["name"].recipe_id is not None:
                        form.cleaned_data["name"].supplier_def_id=Suppliers.objects.get(name="Internal")
                    message="{} {} Has Been Reactivated".format(word, form.cleaned_data["name"].name)
                form.cleaned_data["name"].save()
                messages.success(httprequest, message)
                return HttpResponseRedirect(reverse("stock_web:activreag"))
    else:
        form = form()

    submiturl = reverse("stock_web:activreag")
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Edit Data")

    return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl, "active":"admin"})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def changemin(httprequest, pk):
    submiturl = reverse("stock_web:changemin",args=[pk])
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Edit Data")
    if pk=="_":
        header = ["Select Reagent to Change Minimum Stock Level"]
        form=ChangeMinForm1
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST)
                if form.is_valid():
                    return HttpResponseRedirect(reverse("stock_web:changemin", args=[form.cleaned_data["name"].pk]))
        else:
            form = form()
        return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})
    else:
        item=Reagents.objects.get(pk=int(pk))
        if item.track_vol==True:
            header = ["Select New Default Minimum Stock Level (in µl) for {}".format(item)]
        else:
            header = ["Select New Default Minimum Stock Level for {}".format(item)]
        form=ChangeMinForm
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:changedefsup", args=["_"]))
            else:
                form = form(httprequest.POST, initial = {"old":item.min_count})
                if form.is_valid():
                    item.min_count=form.cleaned_data["number"]
                    item.save()
                    messages.success(httprequest, "Minimum Stock Number for {} has changed to {}{}".format(item,form.cleaned_data["number"], "µl" if item.track_vol==True else ""))
                    return HttpResponseRedirect(reverse("stock_web:listinv"))
        else:
            form = form(initial = {"old":item.min_count})
    return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def changedefsup(httprequest, pk):
    submiturl = reverse("stock_web:changedefsup",args=[pk])
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Edit Data")
    if pk=="_":
        header = ["Select Reagent to Change Default Supplier - THIS WILL NOT AFFECT EXISTING ITEMS"]
        form=ChangeDefSupForm1
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST)
                if form.is_valid():
                    return HttpResponseRedirect(reverse("stock_web:changedefsup", args=[form.cleaned_data["name"].pk]))
        else:
            form = form()
        return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})
    else:
        item=Reagents.objects.get(pk=int(pk))
        header = ["Select New Default Supplier for {} - THIS WILL NOT AFFECT EXISTING ITEMS".format(item)]
        form=ChangeDefSupForm
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:changedefsup", args=["_"]))
            else:
                form = form(httprequest.POST, initial = {"old":item.supplier_def})
                if form.is_valid():
                    item.supplier_def=form.cleaned_data["supplier_def"]
                    item.save()
                    if form.cleaned_data["supplier_def"] is not None:
                        messages.success(httprequest, "Default supplier for {} has changed to {}".format(item,form.cleaned_data["supplier_def"].name))
                    elif form.cleaned_data["supplier_def"] is None:
                        messages.success(httprequest, "Default supplier for {} had been removed".format(item))
                    return HttpResponseRedirect(reverse("stock_web:listinv"))
        else:
            form = form(initial = {"supplier_def":item.supplier_def,
                                   "old":item.supplier_def})
    return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def changedefteam(httprequest, pk):
    submiturl = reverse("stock_web:changedefteam",args=[pk])
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Edit Data")
    if pk=="_":
        header = ["Select Reagent to Change Default Team - THIS WILL NOT AFFECT EXISTING ITEMS"]
        form=ChangeDefTeamForm1
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST)
                if form.is_valid():
                    return HttpResponseRedirect(reverse("stock_web:changedefteam", args=[form.cleaned_data["name"].pk]))
        else:
            form = form()
        return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})
    else:
        item=Reagents.objects.get(pk=int(pk))
        header = ["Select New Default Team for {} - THIS WILL NOT AFFECT EXISTING ITEMS".format(item)]
        form=ChangeDefTeamForm
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:changedefteam", args=["_"]))
            else:
                form = form(httprequest.POST, initial = {"old":item.team_def})
                if form.is_valid():
                    item.team_def=form.cleaned_data["team_def"]
                    item.save()
                    messages.success(httprequest, "Default Team for {} has changed to {}".format(item,form.cleaned_data["team_def"].name))
                    return HttpResponseRedirect(reverse("stock_web:listinv"))
        else:
            form = form(initial = {"team_def":item.team_def,
                                   "old":item.team_def})
    return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def removesup(httprequest):
    header = ["Select Supplier To Remove"]
    form=RemoveSupForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():
                form.cleaned_data["supplier"].delete()
                messages.success(httprequest, "Supplier {} Has Been Deleted".format(form.cleaned_data["supplier"].name))
                return HttpResponseRedirect(reverse("stock_web:listinv"))
    else:
        form = form()
    submiturl = reverse("stock_web:removesup")
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Edit Data")
    return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def editinv(httprequest, pk):
    submiturl = reverse("stock_web:editinv",args=[pk])
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Edit Data")
    if pk=="_":
        title="***WARNING - ONLY TO BE USED TO CORRECT DATA ENTRY ERRORS. IT MAY NOT BE POSSIBLE TO UNDO CHANGES MADE HERE***"
        form=EditInvForm
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or httprequest.POST["submit"] != "search":
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST)
                if form.is_valid():
                    return HttpResponseRedirect(reverse("stock_web:editinv", args=[Inventory.objects.get(internal=Internal.objects.get(batch_number=form.cleaned_data["item"])).pk]))
        else:
            form = form()
        return render(httprequest, "stock_web/undoform.html", {"header":title, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})
    else:
        item = Inventory.objects.get(pk=int(pk))
        if item.reagent.track_vol==True:
##            messages.success(httprequest, "It is currently not possible to undo actions done to cytogenetics reagents using this form. Please contact an Administrator")
##            return HttpResponseRedirect(reverse("stock_web:item",args=[item.pk]))
            return render(httprequest, "stock_web/list_item.html", _vol_context(httprequest, item, "undo"))
        else:
            return render(httprequest, "stock_web/list_item.html", _item_context(httprequest, item, "undo"))

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def undoitem(httprequest, task, pk):
    item=Inventory.objects.get(pk=int(pk))
    submiturl = reverse("stock_web:undoitem",args=[task, pk])
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Edit Data")
    subheading = None
    if task in ["delete", "unopen", "reopen"]:
        form = DeleteForm
        title=["ARE YOU SURE YOU WANT TO {} ITEM {} - {}".format(task.upper(), item.internal, item.reagent)]
        if task=="unopen" and item.reagent.track_vol==True and item.current_vol!=item.vol_rec:
            title+=["THIS WILL REMOVE ALL USES OF THIS REAGENT AND SET ITS VOLUME BACK TO ITS VOLUME RECEIVED"]
        #pdb.set_trace()
        if httprequest.method=="POST":

            if "submit" not in httprequest.POST:
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST)
                if form.is_valid():
                    if form.cleaned_data["sure"]==True:
                        with transaction.atomic():
                            if task=="reopen":
                                item.finished=0
                                item.fin_user=None
                                item.fin_text=None
                                item.reagent.save()
                                item.save()
                            elif task=="unopen":
                                sols=Solutions.objects.filter(Q(comp1=item)|Q(comp2=item)|Q(comp3=item)|
                                                              Q(comp4=item)|Q(comp5=item)|Q(comp6=item)|
                                                              Q(comp7=item)|Q(comp8=item)|Q(comp9=item)|
                                                              Q(comp10=item))
                                if len(sols)!=0:
                                    messages.info(httprequest, "UNABLE TO UNOPEN ITEM AS IT'S USED IN THE FOLLOWING SOLUTION(S)")
                                    for sol in sols:
                                        #pdb.set_trace()
                                        messages.info(httprequest,Inventory.objects.get(sol=sol))
                                    return HttpResponseRedirect(reverse("stock_web:undoitem", args=[task,pk]))
                                item.date_op=None
                                item.is_op=0
                                item.op_user_id=None
                                if item.reagent.track_vol:
                                    item.reagent.count_no+=sum([x.used for x in VolUsage.objects.filter(item=item)])
                                    item.reagent.save()
                                    item.last_usage=None
                                    item.current_vol=item.vol_rec
                                    item.save()
                                    VolUsage.objects.filter(item=item).delete()
                                else:
                                    item.reagent.count_no+=1
                                    item.reagent.save()
                                item.save()
                            if task=="delete":
                                if item.reagent.track_vol==False:
                                    item.reagent.count_no-=1
                                else:
                                    item.reagent.count_no-=item.current_vol
                                item.reagent.save()
                                item.delete()
                                if item.sol is not None:
                                    sol=item.sol
                                    if item.reagent.track_vol==True:
                                        for i in range(sol.recipe.length()):
                                            comp=eval('sol.comp{}'.format(i+1))
                                            uses=VolUsage.objects.filter(item=comp).order_by("id").reverse()
                                            last_use=comp.last_usage
                                            if last_use==uses.get(sol=sol):
                                                try:
                                                    comp.last_usage=uses[1]
                                                except:
                                                    comp.last_usage=None


                                            comp.current_vol+=uses.get(sol=sol).used
                                            comp.reagent.count_no+=uses.get(sol=sol).used
                                            if comp.finished==True:
                                                comp.finished=False
                                                comp.date_fin=None
                                                comp.fin_user=None
                                                comp.fin_text=None
                                            comp.save()
                                            comp.reagent.save()
                                            uses.get(sol=sol).delete()

                                        message="ITEM DELETED AND VOLUMES OF COMPONENTS USED FOR SOLUTION HAVE BEEN REPLACED"
                                    else:
                                        message='Item {} has been deleted!'.format(item)
                                    sol.delete()
                                else:
                                    message='Item {} has been deleted!'.format(item)
                                messages.success(httprequest,message)
                                return HttpResponseRedirect(reverse("stock_web:listinv"))



                    return HttpResponseRedirect(reverse("stock_web:editinv", args=[pk]))
        else:
            form = form()
    elif task=="unval":
        form=UnValForm
        title=["ARE YOU SURE YOU WANT TO UN-VALIDATE ITEM {} - {}".format(item.internal, item.reagent)]
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST)
                form.fields["all_type"].choices = [(0,"NO"),
                                          (1,"YES - Only For This Reagent"),
                                          (2,"YES - All Items On {}, Regardless of Reagent".format(item.val))]

                if form.is_valid():
                    if form.cleaned_data["sure"]==True:
                        current_val_id=item.val_id
                        items=[]
                        if int(form.cleaned_data["all_type"])==0:
                            item.val_id=None
                            item.save()
                        elif int(form.cleaned_data["all_type"])==1:
                            items=Inventory.objects.filter(val_id=current_val_id, reagent_id=item.reagent_id)
                            messages.success(httprequest, "UNVALIDATED: {}".format(" \n".join(test.internal.batch_number for test in items)))
                            items.update(val_id=None)
                        elif int(form.cleaned_data["all_type"])==2:
                            items=Inventory.objects.filter(val_id=current_val_id)
                            messages.success(httprequest, "UNVALIDATED: {}".format(" \n".join(test.internal.batch_number for test in items)))
                            items.update(val_id=None)
                        if len(Inventory.objects.filter(val_id=current_val_id))==0:
                            Validation.objects.get(pk=current_val_id).delete()
                    return HttpResponseRedirect(reverse("stock_web:editinv", args=[pk]))
        else:
            #pdb.set_trace()
            form = form()
            form.fields["all_type"].choices = [(0,"NO"),
                                          (1,"YES - Only For This Reagent"),
                                          (2,"YES - All Items On {}, Regardless of Reagent".format(item.val))]
    elif task=="unuse":
        form=ChangeUseForm
        title=["UNUSE {} - {}".format(item.internal, item.reagent)]
        subheading = ["Enter the a new volume used.",
                      "0 Deletes the usage record.",
                      "Entering a negative number will increase the volume left"]
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST)
                if form.is_valid():
                    if form.cleaned_data["sure"]==True:
                        uses=VolUsage.objects.filter(item=item).order_by("id").reverse()
                        use=item.last_usage
                        if use.sol is not None:
                            messages.success(httprequest, "You cannot edit this usage as it was part of solution {}.".format(Inventory.objects.get(sol=use.sol)))
                            return HttpResponseRedirect(reverse("stock_web:undoitem", args=[task,pk]))
                        item.current_vol+=(use.used-int(form.cleaned_data["vol_used"]))
                        item.reagent.count_no+=(use.used-int(form.cleaned_data["vol_used"]))
                        item.reagent.save()
                        if item.finished==True:
                            item.finished=False
                            item.date_fin=None
                            item.fin_user=None
                            item.fin_text=None
                        if int(form.cleaned_data["vol_used"])==0:
                            if len(uses)>1:
                                item.last_usage=uses[1]
                            else:
                                item.last_usage=None
                            item.save()
                            use.delete()
                        else:
                            use.end=use.start-int(form.cleaned_data["vol_used"])
                            use.used=int(form.cleaned_data["vol_used"])
                            use.user=httprequest.user
                            use.save()
                            item.save()
                        return HttpResponseRedirect(reverse("stock_web:editinv", args=[pk]))
        else:
            #pdb.set_trace()
            form = form(initial = {"vol_used":item.last_usage.used,
                                   "last_usage":item.last_usage.used,
                                   "current_vol":item.current_vol})
    return render(httprequest, "stock_web/form.html", {"header":title, "subheading":subheading, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})
